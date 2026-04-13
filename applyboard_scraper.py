"""
ApplyBoard Program Scraper
Run on your LOCAL machine.

Usage:
    python applyboard_scraper.py              # interactive (prompts for filters)
    python applyboard_scraper.py --debug      # saves raw HTML files for inspection
    python applyboard_scraper.py --no-details # skip detail pages (faster)
    python applyboard_scraper.py --headless   # headless browser
    python applyboard_scraper.py --login      # prompts for ApplyBoard login
"""

import argparse
import asyncio
from getpass import getpass
import os
import random
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import pandas as pd
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, Page

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────

BASE_URL    = "https://www.applyboard.com/search"
DETAIL_BASE = "https://www.applyboard.com"
LOGIN_URL   = "https://accounts.applyboard.com/oauth2/default/v1/authorize?client_id=0oasbh5xhhoozpCwp5d6&redirect_uri=https%3A%2F%2Fwww.applyboard.com%2Fusers%2Fauth%2Foktaoauth%2Fcallback&response_type=code&scope=openid+profile+email+offline_access&state=b77431ea4c500dd33433be2db2ba2a468e60d695d666b30e"
AGENT_URL   = "https://www.applyboard.com/agent"

# These match the exact flag params in ApplyBoard's search URL
DEFAULT_FLAGS = {
    "filter[conditional_offer]":        "f",
    "filter[free_application_only]":    "f",
    "filter[exclude_visa_cap_programs]":"f",
    "filter[pgwp_available]":           "f",
    "filter[ignore_availability]":      "f",
    "filter[unpublished]":              "f",
    "filter[include_pathways]":         "t",
    "sort":                             "-success_score",
}

PAGE_SIZE      = 48
PAGE_LOAD_WAIT = 10_000   # ms
FILTER_WAIT    = 20_000   # ms
OPTION_WAIT    = 15_000   # ms
STEALTH_MIN    = 1.5
STEALTH_MAX    = 3.2
SCROLL_STEPS   = 8
SCROLL_PAUSE   = 800
LOGIN_WAIT     = 30_000
DEFAULT_TIMEOUT = 45_000
SESSION_STATE_PATH = Path("applyboard_session.json")

SYSTEM_BROWSER_PATHS = [
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
]

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.122 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
]

MONTH_NAME_LOOKUP = {
    "jan": "January",
    "january": "January",
    "feb": "February",
    "february": "February",
    "mar": "March",
    "march": "March",
    "apr": "April",
    "april": "April",
    "may": "May",
    "jun": "June",
    "june": "June",
    "jul": "July",
    "july": "July",
    "aug": "August",
    "august": "August",
    "sep": "September",
    "sept": "September",
    "september": "September",
    "oct": "October",
    "october": "October",
    "nov": "November",
    "november": "November",
    "dec": "December",
    "december": "December",
}

MONTH_TOKEN_PATTERN = (
    r"Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|"
    r"Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?"
)

INTAKE_SECTION_STOP_LABELS = [
    "Scholarships",
    "Similar Programs",
    "Post-Study Work Visa",
    "Application Fee",
    "Other Fees",
    "Admission Requirements",
    "Academic Background",
    "Minimum Language Test Scores",
]

# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def rand_sleep(lo=STEALTH_MIN, hi=STEALTH_MAX):
    time.sleep(random.uniform(lo, hi))


async def human_pause(lo_ms: int = 900, hi_ms: int = 1800) -> None:
    await asyncio.sleep(random.uniform(lo_ms, hi_ms) / 1000)


def has_saved_session() -> bool:
    return SESSION_STATE_PATH.exists() and SESSION_STATE_PATH.is_file()

def clean(text) -> str:
    value = (text or "").strip()
    replacements = {
        "Â£": "£",
        "Ł": "£",
        "Â€": "€",
        "Â$": "$",
        "Â ": " ",
        "Â": "",
        "â€™": "'",
        "â€“": "-",
        "â€”": "-",
        "â€˜": "'",
        "â€œ": '"',
        "â€\x9d": '"',
        "\xa0": " ",
    }
    for bad, good in replacements.items():
        value = value.replace(bad, good)
    return re.sub(r'\s+', ' ', value)


def strip_ui_noise(text: str) -> str:
    value = clean(text)
    value = re.sub(r'\(?\bopen in new tab\b\)?', '', value, flags=re.I)
    value = re.sub(r'\s{2,}', ' ', value).strip(" -,:|()")
    return clean(value)


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", strip_ui_noise(text).lower()).strip()


def save_debug_html(html: str, label: str = "page"):
    path = Path(f"debug_{label}_{int(time.time())}.html")
    path.write_text(html, encoding="utf-8")
    print(f"   💾  Debug HTML → {path}")


def mask_email(email: str) -> str:
    if not email or "@" not in email:
        return "***"
    name, domain = email.split("@", 1)
    visible = name[:2] if len(name) > 2 else name[:1]
    return f"{visible}***@{domain}"


def sanitize_filename(value: str, fallback: str = "item") -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", clean(value)).strip("._-")
    return cleaned[:80] or fallback


def get_system_browser_path() -> str | None:
    for browser_path in SYSTEM_BROWSER_PATHS:
        if Path(browser_path).exists():
            return browser_path
    return None


async def wait_for_first_visible(page: Page, selectors: list[str], timeout_ms: int):
    deadline = time.monotonic() + (timeout_ms / 1000)
    while time.monotonic() < deadline:
        for sel in selectors:
            try:
                loc = page.locator(sel).first
                if await loc.count() > 0 and await loc.is_visible():
                    return loc
            except Exception:
                continue
        await page.wait_for_timeout(300)
    return None


async def click_first_visible(page: Page, selectors: list[str], timeout_ms: int = 5_000) -> bool:
    deadline = time.monotonic() + (timeout_ms / 1000)
    while time.monotonic() < deadline:
        for sel in selectors:
            try:
                locator = page.locator(sel).first
                if await locator.count() > 0 and await locator.is_visible():
                    await locator.click()
                    return True
            except Exception:
                continue
        await page.wait_for_timeout(250)
    return False


async def settle_page(page: Page, pause_ms: int = 1500) -> None:
    for state in ("domcontentloaded", "load"):
        try:
            await page.wait_for_load_state(state, timeout=10_000)
        except Exception:
            pass
    try:
        await page.wait_for_load_state("networkidle", timeout=5_000)
    except Exception:
        pass
    await human_pause(max(500, pause_ms - 600), pause_ms + 400)


def is_logged_in_portal_page(page_url: str, body_text: str) -> bool:
    current_url = (page_url or "").lower()
    text = clean(body_text)
    lowered = text.lower()

    if not current_url.startswith("https://www.applyboard.com/"):
        return False
    if "accounts.applyboard.com" in current_url or "unauthorized" in current_url:
        return False
    if re.search(r"\blog in\b", text, re.I) and re.search(r"\bregister\b", text, re.I):
        return False

    portal_markers = [
        "search",
        "applications",
        "students",
        "recruitment partners",
        "offers",
        "commissions",
        "deals",
        "messages",
        "profile",
        "programs",
        "schools",
        "agent",
        "dashboard",
    ]
    if any(marker in lowered for marker in portal_markers):
        return True

    return len(text) >= 80


async def read_locator_value(locator) -> str:
    try:
        value = await locator.input_value()
        if value:
            return clean(value)
    except Exception:
        pass

    try:
        value = await locator.get_attribute("value")
        if value:
            return clean(value)
    except Exception:
        pass

    try:
        text = await locator.inner_text()
        if text:
            return clean(text)
    except Exception:
        pass

    return ""


async def open_institution_filter(page: Page) -> None:
    """
    Open the school filter area when ApplyBoard keeps it collapsed under
    sections like Schools or All filters.
    """
    openers = [
        'button:has-text("Institution (School)")',
        'button:has-text("Institution")',
        'button:has-text("School")',
        'button:has-text("Schools")',
        '[role="button"]:has-text("Institution (School)")',
        '[role="button"]:has-text("Institution")',
        '[role="button"]:has-text("School")',
        '[role="button"]:has-text("Schools")',
        'summary:has-text("Institution (School)")',
        'summary:has-text("Institution")',
        'summary:has-text("School")',
        'summary:has-text("Schools")',
        'button:has-text("All filters")',
        '[role="button"]:has-text("All filters")',
        'button:has-text("More filters")',
        '[role="button"]:has-text("More filters")',
        'button[aria-label*="filter" i]',
        '[role="button"][aria-label*="filter" i]',
    ]

    # Try likely openers a few times because one click may reveal another layer.
    for _ in range(3):
        clicked = await click_first_visible(page, openers, timeout_ms=2_000)
        if not clicked:
            break
        await page.wait_for_timeout(900)


async def read_selected_university(page: Page) -> str:
    """
    Many filter widgets clear the search box after selection, so look for the
    chosen school in checked options, chips, or selected filter labels.
    """
    selectors = [
        'label:has(input[type="checkbox"]:checked)',
        '[role="option"][aria-selected="true"]',
        '[aria-checked="true"]',
        '[data-state="checked"]',
        '[class*="selected" i]',
        '[class*="chip" i]',
        '[class*="tag" i]',
    ]

    for selector in selectors:
        try:
            locator = page.locator(selector)
            count = await locator.count()
            for index in range(min(count, 10)):
                text = clean(await read_locator_value(locator.nth(index)))
                if text:
                    return text
        except Exception:
            continue

    return ""

def build_search_url(page_number: int = 1, base_url: str = None,
                     country: str = None, subject: str = None,
                     degree_level: str = None) -> str:
    """
    Build the exact URL format ApplyBoard uses, e.g.:
    /search?filter[school_ids]=1715&page[number]=1&page[size]=48&sort=-success_score...
    """
    params = dict(DEFAULT_FLAGS)
    if base_url:
        parsed = urlsplit(base_url)
        params.update(dict(parse_qsl(parsed.query, keep_blank_values=True)))
    params["page[number]"] = str(page_number)
    params["page[size]"]   = str(PAGE_SIZE)

    if country:
        params["filter[country]"] = country
    if subject:
        params["filter[subject]"] = subject
    if degree_level:
        params["filter[degree_type]"] = degree_level

    return urlunsplit(("https", "www.applyboard.com", "/search", urlencode(params), ""))


def with_school_id(base_url: str, school_id: str) -> str:
    parsed = urlsplit(base_url or BASE_URL)
    params = dict(parse_qsl(parsed.query, keep_blank_values=True))
    params.update(DEFAULT_FLAGS)
    params["filter[school_ids]"] = school_id
    params["page[number]"] = params.get("page[number]", "1")
    params["page[size]"] = params.get("page[size]", str(PAGE_SIZE))
    return urlunsplit(("https", "www.applyboard.com", "/search", urlencode(params), ""))


def has_university_filter_param(url: str) -> bool:
    return any(
        token in (url or "")
        for token in (
            "filter%5Bschool_ids%5D=",
            "filter[school_ids]=",
            "filter%5Bschool_group_ids%5D=",
            "filter[school_group_ids]=",
        )
    )

# ──────────────────────────────────────────────────────────────────────────────
# BROWSER
# ──────────────────────────────────────────────────────────────────────────────

async def create_browser(playwright, headless: bool = False, storage_state_path: Path | None = None):
    launch_kwargs = {
        "headless": headless,
        "slow_mo": 85,
        "args": [
            "--no-sandbox",
            "--disable-dev-shm-usage",
        ],
    }
    try:
        browser = await playwright.chromium.launch(**launch_kwargs)
    except Exception as exc:
        browser_path = get_system_browser_path()
        if not browser_path:
            raise RuntimeError(
                "Could not launch Playwright Chromium, and no system Chrome/Edge browser "
                "was found. Install Chromium with 'python -m playwright install chromium'."
            ) from exc
        print(f"   Using system browser executable -> {browser_path}")
        browser = await playwright.chromium.launch(
            executable_path=browser_path,
            **launch_kwargs,
        )
    context_kwargs = {
        "user_agent": random.choice(USER_AGENTS),
        "viewport": {"width": 1440, "height": 900},
        "locale": "en-US",
        "timezone_id": "America/Toronto",
        "extra_http_headers": {
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        },
    }
    if storage_state_path and storage_state_path.exists():
        context_kwargs["storage_state"] = str(storage_state_path)
    ctx = await browser.new_context(**context_kwargs)
    ctx.set_default_timeout(DEFAULT_TIMEOUT)
    ctx.set_default_navigation_timeout(DEFAULT_TIMEOUT)
    return browser, ctx


async def login_to_applyboard(page: Page, email: str, password: str, debug: bool = False) -> None:
    if not email or not password:
        raise ValueError("ApplyBoard login requires both an email and password.")

    print(f"\n[login] Logging in to ApplyBoard as {mask_email(email)}")
    await page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=60_000)
    await page.wait_for_timeout(2_500)

    email_selectors = [
        'input[name="identifier"]',
        'input[name="username"]',
        'input[type="email"]',
        'input[autocomplete="username"]',
        'input[placeholder*="email" i]',
        'input[placeholder*="username" i]',
    ]
    password_selectors = [
        'input[name="credentials.passcode"]',
        'input[name="password"]',
        'input[type="password"]',
        'input[autocomplete="current-password"]',
        'input[placeholder*="password" i]',
    ]
    submit_selectors = [
        'input[type="submit"]',
        'button:has-text("Log In")',
        'button:has-text("Sign In")',
        'button:has-text("Next")',
        'button:has-text("Continue")',
        '[role="button"]:has-text("Log In")',
        '[role="button"]:has-text("Sign In")',
    ]

    email_input = await wait_for_first_visible(page, email_selectors, LOGIN_WAIT)
    if not email_input:
        if debug:
            save_debug_html(await page.content(), "login_email_not_found")
        raise RuntimeError("Could not find the ApplyBoard email input on the login page.")

    await email_input.click()
    await email_input.fill("")
    await email_input.type(email, delay=50)
    await page.wait_for_timeout(500)

    password_input = await wait_for_first_visible(page, password_selectors, 2_500)
    if not password_input:
        await click_first_visible(page, submit_selectors, timeout_ms=4_000)
        password_input = await wait_for_first_visible(page, password_selectors, LOGIN_WAIT)

    if not password_input:
        if debug:
            save_debug_html(await page.content(), "login_password_not_found")
        raise RuntimeError(
            "The password field did not appear after entering the email. "
            "ApplyBoard may have changed the login flow."
        )

    await password_input.click()
    await password_input.fill(password)
    await page.wait_for_timeout(400)

    if not await click_first_visible(page, submit_selectors, timeout_ms=5_000):
        try:
            await password_input.press("Enter")
        except Exception:
            if debug:
                save_debug_html(await page.content(), "login_submit_not_found")
            raise RuntimeError("Could not find the ApplyBoard login submit button.")

    try:
        await page.wait_for_url(lambda url: "accounts.applyboard.com" not in url, timeout=LOGIN_WAIT)
    except Exception:
        await page.wait_for_timeout(3_000)

    if "accounts.applyboard.com" in page.url:
        page_text = clean(await page.locator("body").inner_text())
        if debug:
            save_debug_html(await page.content(), "login_not_completed")
        if re.search(r"(verify|authenticator|multi-factor|security key|one-time passcode)", page_text, re.I):
            raise RuntimeError(
                "Login reached an additional verification step (for example MFA). "
                "Complete that step manually, then rerun the scraper."
            )
        if re.search(r"(incorrect|invalid|required|unable to sign in)", page_text, re.I):
            raise RuntimeError("ApplyBoard rejected the login. Check the email/password and try again.")
        raise RuntimeError(f"Login did not complete successfully. Current page: {page.url}")

    await page.goto(AGENT_URL, wait_until="domcontentloaded", timeout=60_000)
    await settle_page(page, pause_ms=2_500)

    agent_text = clean(await page.locator("body").inner_text())
    if not is_logged_in_portal_page(page.url, agent_text):
        if debug:
            save_debug_html(await page.content(), "agent_portal_not_loaded")
        raise RuntimeError(
            f"Login succeeded, but the ApplyBoard agent portal did not load as expected. Current page: {page.url}"
        )

    print(f"   [ok] Logged in successfully -> {page.url}")


async def has_active_agent_session(page: Page, debug: bool = False) -> bool:
    try:
        await page.goto(AGENT_URL, wait_until="domcontentloaded", timeout=60_000)
        await settle_page(page, pause_ms=2_200)
    except Exception:
        return False

    current_url = page.url.lower()
    body_text = clean(await page.locator("body").inner_text())
    if not is_logged_in_portal_page(current_url, body_text):
        return False

    if debug:
        print(f"   [session] Active ApplyBoard session detected -> {page.url}")
    return True

# ──────────────────────────────────────────────────────────────────────────────
# FILTERS
# ──────────────────────────────────────────────────────────────────────────────

async def apply_university_filter(page: Page, university_name: str, debug: bool = False) -> str | None:
    if not university_name:
        return page.url

    print(f"\n🔍  Applying university filter: {university_name}")

    await open_institution_filter(page)

    input_selectors = [
        'input[placeholder*="institution" i]',
        'input[placeholder*="university" i]',
        'input[placeholder*="school" i]',
        'input[placeholder*="search school" i]',
        'input[placeholder*="search institutions" i]',
        'input[placeholder*="search" i]',
        'input[aria-label*="institution" i]',
        'input[aria-label*="school" i]',
        'input[aria-label*="search" i]',
        '[role="combobox"][aria-label*="institution" i]',
        '[role="combobox"][aria-label*="school" i]',
        '[role="combobox"][aria-label*="search" i]',
    ]

    institution_input = await wait_for_first_visible(page, input_selectors, FILTER_WAIT)
    if not institution_input:
        await open_institution_filter(page)
        institution_input = await wait_for_first_visible(page, input_selectors, 6_000)

    if not institution_input:
        print("   Waiting for institution filter timed out.")
        print("   ⚠  Could not find Institution (School) input.")
        if debug:
            save_debug_html(await page.content(), "institution_input_not_found")
        return None

    await institution_input.click()
    await page.wait_for_timeout(500)
    await institution_input.fill("")
    await institution_input.type(university_name, delay=80)
    await page.wait_for_timeout(1200)

    partial_tokens = [token for token in re.split(r"\s+", university_name) if len(token) >= 4]
    checkbox_targets = [
        ('label', True),
        ('[role="option"]', False),
        ('[role="listbox"] li', False),
        ('[class*="option" i]', False),
        ('[class*="suggestion" i]', False),
        ('[class*="autocomplete" i] li', False),
    ]

    async def choose_best_candidate(candidates, expected_text: str):
        best_locator = None
        best_score = None
        expected_normalized = normalize_text(expected_text)
        tokens = [token for token in expected_normalized.split() if len(token) >= 3]

        count = await candidates.count()
        for index in range(min(count, 20)):
            try:
                candidate = candidates.nth(index)
                if not await candidate.is_visible():
                    continue

                candidate_text = strip_ui_noise(await read_locator_value(candidate))
                candidate_normalized = normalize_text(candidate_text)
                if not candidate_normalized:
                    continue

                score = 0
                if expected_normalized and expected_normalized in candidate_normalized:
                    score += 50
                score += sum(5 for token in tokens if token in candidate_normalized)
                if "all campuses" in candidate_normalized:
                    score += 100
                if "campus" in candidate_normalized:
                    score += 10
                score -= len(candidate_text)

                if best_score is None or score > best_score:
                    best_score = score
                    best_locator = candidate
            except Exception:
                continue

        return best_locator

    def extract_numeric_identifier(value: str, allow_generic: bool = True) -> str:
        if not value:
            return ""

        patterns = [r'group_(\d+)']
        if allow_generic:
            patterns.append(r'(?<!\d)(\d{3,})(?!\d)')

        for pattern in patterns:
            match = re.search(pattern, value, re.I)
            if match:
                return match.group(1)
        return ""

    async def extract_school_id_from_candidate(candidate) -> str:
        attr_names = ["value", "data-value", "data-id", "id", "for", "aria-describedby"]
        locators = [
            candidate,
            candidate.locator('input[type="checkbox"]').first,
            candidate.locator('input').first,
            candidate.locator('[data-id]').first,
            candidate.locator('[data-value]').first,
            candidate.locator('xpath=..').first,
            candidate.locator('xpath=../following-sibling::*[1]').first,
            candidate.locator('xpath=../..').first,
        ]

        for locator in locators:
            try:
                if await locator.count() == 0:
                    continue
                for attr_name in attr_names:
                    attr_value = await locator.get_attribute(attr_name)
                    if not attr_value:
                        continue
                    extracted = extract_numeric_identifier(attr_value)
                    if extracted:
                        return extracted

                try:
                    locator_html = await locator.evaluate("(el) => el.outerHTML")
                except Exception:
                    locator_html = ""
                extracted = extract_numeric_identifier(locator_html, allow_generic=False)
                if extracted:
                    return extracted
            except Exception:
                continue

        return ""

    async def apply_candidate_selection(candidate, candidate_text: str, prefer_checkbox: bool) -> None:
        checkbox = candidate.locator('input[type="checkbox"]').first
        if await checkbox.count() > 0 and await checkbox.is_visible():
            if not await checkbox.is_checked():
                await checkbox.check()
            else:
                await candidate.click(force=True)
            return

        # Some ApplyBoard filter widgets use a role=option item plus a sibling hit area.
        alias_locators = []
        if candidate_text:
            try:
                escaped = candidate_text.replace('"', '\\"')
                alias_locators.extend([
                    page.locator(f'[aria-label="{escaped}"]').first,
                    page.locator(f'[aria-label="{escaped}"]').last,
                ])
            except Exception:
                pass

        alias_locators.extend([
            candidate.locator('xpath=..').first,
            candidate.locator('xpath=../following-sibling::*[1]').first,
        ])

        for locator in [candidate, *alias_locators]:
            try:
                if await locator.count() == 0 or not await locator.is_visible():
                    continue
                await locator.click(force=True)
                return
            except Exception:
                continue

        try:
            await candidate.focus()
            await page.keyboard.press("Enter")
        except Exception:
            pass

    deadline = time.monotonic() + (OPTION_WAIT / 1000)
    while time.monotonic() < deadline:
        for selector, prefer_checkbox in checkbox_targets:
            texts = [university_name, *partial_tokens[:3]]
            for text in texts:
                try:
                    candidates = page.locator(selector).filter(has_text=text)
                    if await candidates.count() == 0:
                        continue

                    candidate = await choose_best_candidate(candidates, university_name)
                    if candidate is None:
                        continue

                    if not await candidate.is_visible():
                        continue

                    candidate_text = strip_ui_noise(await read_locator_value(candidate))
                    if candidate_text:
                        print(f"   Selecting institution option: {candidate_text}")

                    candidate_school_id = await extract_school_id_from_candidate(candidate)
                    await apply_candidate_selection(candidate, candidate_text, prefer_checkbox)

                    try:
                        await page.wait_for_load_state("networkidle", timeout=8_000)
                    except Exception:
                        pass
                    await page.wait_for_timeout(1500)

                    if has_university_filter_param(page.url):
                        print(f"   University filter applied → {page.url}")
                        return page.url

                    candidate_normalized = normalize_text(candidate_text)
                    expected_normalized = normalize_text(university_name)
                    selected_value = await read_selected_university(page)
                    selected_normalized = normalize_text(selected_value)
                    if selected_normalized and (
                        expected_normalized in selected_normalized
                        or selected_normalized in expected_normalized
                    ):
                        if has_university_filter_param(page.url):
                            print(f"   University filter applied via selected option state → {page.url}")
                            return page.url
                        if candidate_school_id:
                            fallback_url = with_school_id(page.url, candidate_school_id)
                            print(f"   University filter applied via selected option fallback → {fallback_url}")
                            return fallback_url

                    if candidate_school_id and candidate_normalized and (
                        expected_normalized in candidate_normalized
                        or candidate_normalized in expected_normalized
                        or "all campuses" in candidate_normalized
                    ):
                        fallback_url = with_school_id(page.url, candidate_school_id)
                        print(f"   University filter applied via metadata fallback → {fallback_url}")
                        return fallback_url
                except Exception:
                    continue

        await page.wait_for_timeout(500)

    print("   Waiting for university checkbox options timed out.")
    if debug:
        save_debug_html(await page.content(), "university_filter_not_applied")
    return None


async def wait_for_filters_ready(page: Page, filters: dict, filtered_base_url: str | None,
                                 debug: bool = False) -> bool:
    checkpoint_url = build_search_url(
        page_number=1,
        base_url=filtered_base_url,
        country=filters.get("country"),
        subject=filters.get("subject"),
        degree_level=filters.get("degree_level"),
    )

    print(f"\n⏳  Loading filters checkpoint → {checkpoint_url}")
    await page.goto(checkpoint_url, wait_until="domcontentloaded", timeout=30_000)
    try:
        await page.wait_for_load_state("networkidle", timeout=8_000)
    except Exception:
        pass
    await page.wait_for_timeout(2500)

    university_name = filters.get("university", "")
    if university_name:
        await open_institution_filter(page)
        institution_selectors = [
            'input[placeholder*="institution" i]',
            'input[placeholder*="school" i]',
            'input[placeholder*="search school" i]',
            'input[placeholder*="search institutions" i]',
            'input[placeholder*="search" i]',
            'input[aria-label*="institution" i]',
            'input[aria-label*="school" i]',
            'input[aria-label*="search" i]',
            '[role="combobox"][aria-label*="institution" i]',
            '[role="combobox"][aria-label*="school" i]',
            '[role="combobox"][aria-label*="search" i]',
        ]
        institution_field = await wait_for_first_visible(page, institution_selectors, FILTER_WAIT)
        if not institution_field:
            await open_institution_filter(page)
            institution_field = await wait_for_first_visible(page, institution_selectors, 6_000)

        if not institution_field:
            print("   ⚠  Filters checkpoint failed: Institution (School) field was not visible.")
            if debug:
                save_debug_html(await page.content(), "filters_checkpoint_missing_institution")
            return False

        actual_value = await read_locator_value(institution_field)
        selected_value = await read_selected_university(page)
        verification_value = actual_value or selected_value
        expected_normalized = normalize_text(university_name)
        actual_normalized = normalize_text(verification_value)
        expected_tokens = [token for token in expected_normalized.split() if len(token) >= 3]
        matches = (
            expected_normalized and actual_normalized and (
                expected_normalized in actual_normalized
                or actual_normalized in expected_normalized
                or all(token in actual_normalized for token in expected_tokens[:3])
            )
        )

        if not matches:
            print("   ⚠  Filters checkpoint failed: university value was not loaded into Institution (School).")
            print(f"      Expected: {university_name}")
            print(f"      Actual input:    {actual_value or '[empty]'}")
            print(f"      Actual selected: {selected_value or '[empty]'}")
            if debug:
                save_debug_html(await page.content(), "filters_checkpoint_university_mismatch")
            return False

        if actual_value:
            print(f"   Institution (School) field confirmed: {actual_value}")
        else:
            print(f"   Institution (School) selection confirmed: {selected_value}")

    print("✅  Filters ready")
    return True

# ──────────────────────────────────────────────────────────────────────────────
# SCROLL
# ──────────────────────────────────────────────────────────────────────────────

async def human_scroll(page: Page):
    height = await page.evaluate("document.body.scrollHeight")
    step = max(height // SCROLL_STEPS, 300)
    for i in range(1, SCROLL_STEPS + 1):
        await page.evaluate(f"window.scrollTo(0, {step * i})")
        await page.wait_for_timeout(SCROLL_PAUSE + random.randint(-150, 200))
    await page.evaluate("window.scrollTo(0, 300)")
    await page.wait_for_timeout(400)

# ──────────────────────────────────────────────────────────────────────────────
# CARD DETECTION + PARSING
# ──────────────────────────────────────────────────────────────────────────────

def detect_cards(soup: BeautifulSoup) -> list[Tag]:
    specific = [
        '[data-testid*="program"]', '[data-testid*="card"]',
        '[class*="ProgramCard"]',   '[class*="program-card"]',
        '[class*="programCard"]',   '[class*="SearchResult"]',
        '[class*="search-result"]', '[class*="ResultCard"]',
        '[class*="result-card"]',   'article',
        '[class*="Program"][class*="Item"]',
    ]
    for sel in specific:
        cards = soup.select(sel)
        if len(cards) >= 2:
            print(f"   🎯  Selector '{sel}' → {len(cards)} cards")
            return cards

    # Structural heuristic fallback
    candidates = []
    for tag in soup.find_all(['div','li','section','article'], recursive=True):
        if (tag.find(['h1','h2','h3','h4','h5']) and
                tag.find('a', href=True) and
                60 < len(tag.get_text(strip=True)) < 2500):
            candidates.append(tag)

    # De-nest: remove elements that are ancestors of other candidates
    cand_ids = {id(c) for c in candidates}
    unique = [c for c in candidates
              if not any(id(p) in cand_ids for p in c.parents)]

    if unique:
        print(f"   🔍  Heuristic → {len(unique)} cards")
        return unique[:200]

    print("   ⚠  No cards detected. Run with --debug to inspect HTML.")
    return []

INSTITUTION_KEYWORDS = (
    "university", "college", "institute", "school", "polytechnic",
    "academy", "faculty", "campus",
)

FIELD_LABELS = {
    "university": ["school", "institution", "university", "college"],
    "country": ["country", "nation"],
    "city": ["city", "campus"],
    "degree_level": ["degree", "level", "credential", "qualification"],
    "subject": ["subject", "discipline", "field", "area"],
    "duration": ["duration", "length"],
    "tuition": ["tuition", "fee", "cost", "price"],
    "language": ["language", "instruction"],
    "intake": ["intake", "start", "semester", "term"],
}


def looks_like_institution(text: str) -> bool:
    normalized = normalize_text(text)
    return any(keyword in normalized for keyword in INSTITUTION_KEYWORDS)


def normalize_institution_name(value: str) -> str:
    cleaned = clean(value)
    if " - " not in cleaned:
        return cleaned
    left, right = [clean(part) for part in cleaned.split(" - ", 1)]
    if looks_like_institution(left) and right and not looks_like_institution(right):
        return left
    return cleaned


def looks_like_location_line(value: str) -> bool:
    cleaned = clean(value)
    if not cleaned:
        return False
    if re.match(r"^[A-Za-z .'-]+,\s*[A-Za-z .'-]+,\s*[A-Z]{2,3}$", cleaned):
        return True
    if re.match(r"^[A-Za-z .'-]+,\s*[A-Za-z .'-]+$", cleaned) and len(cleaned) <= 60:
        return True
    return False


def looks_like_programme_title(value: str) -> bool:
    cleaned = clean(value)
    normalized = normalize_text(cleaned)
    if not normalized or len(cleaned) < 6:
        return False
    if looks_like_institution(cleaned) or looks_like_location_line(cleaned):
        return False
    if re.fullmatch(r"\d+", cleaned):
        return False
    if re.search(r"^(home|overview|admission requirements|scholarships|similar programs|view photos|show more)$", cleaned, re.I):
        return False
    if re.search(r"^(open|likely open|closed|instant submission|high job demand|scholarships available|prime|incentivized|popular|fast acceptance)$", cleaned, re.I):
        return False
    if re.search(r"\b(bachelor|master|msc|mba|phd|diploma|certificate|degree|law|arts|science|business|engineering|management|program|programme)\b", normalized, re.I):
        return True
    return len(cleaned) >= 18 and (" - " in cleaned or len(cleaned.split()) >= 3)


def get_card_chunks(card: Tag, max_len: int = 120) -> list[str]:
    chunks = []
    seen = set()
    for chunk in card.stripped_strings:
        value = clean(chunk)
        if not value or len(value) > max_len:
            continue
        lowered = value.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        chunks.append(value)
    return chunks


def find_chunk_labeled_value(chunks: list[str], labels: list[str], max_len: int = 180) -> str:
    for chunk in chunks:
        for label in labels:
            pattern = rf'^{re.escape(label)}[:\s-]+(.+)$'
            match = re.match(pattern, chunk, re.I)
            if match:
                return clean(match.group(1))[:max_len]
    return ""


def find_labeled_value(text: str, labels: list[str], max_len: int = 180) -> str:
    for label in labels:
        pattern = rf'(?:^|[\s|•]){re.escape(label)}[:\s-]+([^\n|•<>:]{{2,{max_len}}})'
        match = re.search(pattern, text, re.I)
        if match:
            return clean(match.group(1))[:max_len]
    return ""


def is_meaningful_field_value(field_name: str, value: str, labels: list[str]) -> bool:
    normalized = normalize_text(value)
    if not normalized:
        return False

    normalized_labels = {normalize_text(label) for label in labels}
    if normalized in normalized_labels:
        return False

    if field_name == "city" and normalized in {"city", "campus", "campus city", "open in new tab"}:
        return False

    if field_name == "tuition":
        if normalized in {"1st year", "first year", "tuition 1st year", "tuition first year"}:
            return False
        if not re.search(r'(\d|[$£€]|[A-Z]{3})', value):
            return False

    return True


def extract_tuition_value(text: str, chunks: list[str]) -> str:
    tuition_labels = [
        "tuition (1st year)",
        "tuition 1st year",
        "tuition first year",
        "1st year tuition",
        "first year tuition",
        "annual tuition",
        "tuition",
    ]

    for chunk in chunks:
        normalized_chunk = normalize_text(chunk)
        if any(normalize_text(label) in normalized_chunk for label in tuition_labels):
            money_match = re.search(r'([A-Z]{3}|[$£€]|CAD|USD|GBP|EUR)\s*[\d,]+(?:\.\d+)?', chunk)
            if money_match:
                return clean(money_match.group(0))
            value = find_chunk_labeled_value([chunk], tuition_labels, max_len=140)
            if value and is_meaningful_field_value("tuition", value, tuition_labels):
                return value

    patterns = [
        r'tuition\s*\(?(?:1st|first)\s*year\)?[:\s-]+((?:[A-Z]{3}|[$£€]|CAD|USD|GBP|EUR)?\s*[\d,]+(?:\.\d+)?)',
        r'(?:annual\s+)?tuition[:\s-]+((?:[A-Z]{3}|[$£€]|CAD|USD|GBP|EUR)?\s*[\d,]+(?:\.\d+)?)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return clean(match.group(1))

    return ""


def extract_city_value(text: str, chunks: list[str]) -> str:
    city_labels = ["campus city", "city", "campus"]

    for index, chunk in enumerate(chunks):
        cleaned_chunk = strip_ui_noise(chunk)
        normalized_chunk = normalize_text(cleaned_chunk)
        if normalized_chunk == "campus city":
            for next_chunk in chunks[index + 1:index + 4]:
                candidate = strip_ui_noise(next_chunk)
                if candidate and is_meaningful_field_value("city", candidate, city_labels):
                    return candidate

        inline_value = find_chunk_labeled_value([cleaned_chunk], city_labels, max_len=120)
        inline_value = strip_ui_noise(inline_value)
        if inline_value and is_meaningful_field_value("city", inline_value, city_labels):
            return inline_value

    value = strip_ui_noise(find_labeled_value(text, city_labels, max_len=120))
    if value and is_meaningful_field_value("city", value, city_labels):
        return value
    return ""


def normalize_campus_city(value: str) -> str:
    cleaned = clean(value)
    if not cleaned:
        return ""
    first_part = clean(cleaned.split(",", 1)[0])
    if first_part:
        return first_part
    return cleaned


def extract_duration_value(text: str, chunks: list[str]) -> str:
    duration_patterns = [
        r'\b\d+(?:\.\d+)?\s*(?:years?|yrs?|months?|weeks?)\b',
        r'\b(?:one|two|three|four|five|six|seven|eight|nine|ten|eleven|twelve)\s+(?:year|month|week)s?\b',
    ]
    stop_tokens = {"success", "rate", "job", "demand", "scholarship", "available", "open in new tab"}

    def clean_duration_candidate(candidate: str) -> str:
        value = clean(candidate)
        for pattern in duration_patterns:
            match = re.search(pattern, value, re.I)
            if match:
                return clean(match.group(0))
        return ""

    duration_labels = FIELD_LABELS["duration"]

    for chunk in chunks:
        inline_value = find_chunk_labeled_value([chunk], duration_labels, max_len=120)
        if inline_value:
            parsed = clean_duration_candidate(inline_value)
            if parsed:
                return parsed

        lowered = normalize_text(chunk)
        if any(label in lowered for label in duration_labels):
            parsed = clean_duration_candidate(chunk)
            if parsed:
                return parsed

        if not any(token in lowered for token in stop_tokens):
            parsed = clean_duration_candidate(chunk)
            if parsed:
                return parsed

    text_value = find_labeled_value(text, duration_labels, max_len=120)
    parsed = clean_duration_candidate(text_value)
    if parsed:
        return parsed

    for pattern in duration_patterns:
        match = re.search(pattern, text, re.I)
        if match:
            return clean(match.group(0))

    return ""


def extract_month_year_values(text: str) -> list[str]:
    matches = re.findall(rf'\b({MONTH_TOKEN_PATTERN})\s+(20\d{{2}})\b', text or "", re.I)
    values = []
    for month_token, year in matches:
        month_name = MONTH_NAME_LOOKUP.get(month_token.lower())
        if not month_name:
            continue
        value = f"{month_name} {year}"
        if value not in values:
            values.append(value)
    return values


def normalize_intake_value(value: str) -> str:
    intakes = extract_month_year_values(value)
    if intakes:
        return ", ".join(intakes)
    return clean(value)


def extract_detail_header_fields(lines: list[str]) -> tuple[str, str]:
    header_lines = []
    stop_labels = {
        "overview",
        "admission requirements",
        "scholarships",
        "similar programs",
        "program summary",
    }

    for line in lines[:30]:
        cleaned = strip_ui_noise(line)
        normalized = normalize_text(cleaned)
        if not cleaned:
            continue
        if normalized in stop_labels:
            break
        if re.match(r"^you have \d+ important updates$", cleaned, re.I):
            continue
        if re.match(r"^(home|view photos|show more)$", cleaned, re.I):
            continue
        if re.match(r"^(open|likely open|closed|instant submission|high job demand|scholarships available|prime|incentivized|popular|fast acceptance)$", cleaned, re.I):
            continue
        if re.fullmatch(r"\d+", cleaned):
            continue
        header_lines.append(cleaned)

    institution = ""
    programme_name = ""

    for index, line in enumerate(header_lines):
        if looks_like_institution(line):
            institution = normalize_institution_name(line)
            for next_line in header_lines[index + 1:]:
                if looks_like_programme_title(next_line):
                    programme_name = clean(next_line)
                    return institution, programme_name
            break

    if not programme_name:
        for line in header_lines:
            if looks_like_programme_title(line):
                programme_name = clean(line)
                break

    return institution, programme_name


def extract_available_intakes(text: str, chunks: list[str], lines: list[str] | None = None) -> str:
    if lines:
        for section_label in ("Program Intakes", "Available Intakes"):
            section = extract_section_lines(
                lines,
                section_label,
                stop_labels=INTAKE_SECTION_STOP_LABELS,
            )
            if section:
                section_values = extract_month_year_values(" ".join(section))
                if section_values:
                    return ", ".join(section_values)

    intakes = []

    for index, chunk in enumerate(chunks):
        if normalize_text(chunk) in {"available intakes", "program intakes"}:
            for next_chunk in chunks[index + 1:index + 8]:
                matches = extract_month_year_values(next_chunk)
                for value in matches:
                    if value not in intakes:
                        intakes.append(value)
            if intakes:
                return ", ".join(intakes)

    for pattern in (r'available\s+intakes(.{0,300})', r'program\s+intakes(.{0,300})'):
        available_block = re.search(pattern, text, re.I)
        if not available_block:
            continue
        for value in extract_month_year_values(available_block.group(1)):
            if value not in intakes:
                intakes.append(value)

    if intakes:
        return ", ".join(intakes)

    return ""


def choose_program_name(card: Tag, href: str, text: str, chunks: list[str]) -> str:
    candidates = []

    if href:
        for anchor in card.select('a[href*="/programs/"], a[href*="/apply"]'):
            value = clean(anchor.get_text(" ", strip=True))
            if value:
                candidates.append(value)

    for heading in card.find_all(['h1', 'h2', 'h3', 'h4', 'h5']):
        value = clean(heading.get_text(" ", strip=True))
        if value:
            candidates.append(value)

    for element in card.select(
        '[data-testid*="title"], [data-testid*="program"], [data-testid*="name"], '
        '[class*="title" i], [class*="program" i], [class*="name" i]'
    ):
        value = clean(element.get_text(" ", strip=True))
        if value:
            candidates.append(value)

    labeled = find_chunk_labeled_value(chunks, ["program", "title", "name"], max_len=140) or \
        find_labeled_value(text, ["program", "title", "name"], max_len=140)
    if labeled:
        candidates.append(labeled)

    candidates.extend(chunks[:8])

    seen = set()
    fallback = []
    for candidate in candidates:
        candidate = clean(candidate)
        normalized = normalize_text(candidate)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        if looks_like_programme_title(candidate):
            return candidate[:180]
        if looks_like_institution(candidate):
            continue
        if len(candidate) < 4:
            continue
        if looks_like_location_line(candidate):
            continue
        if re.search(r'^(country|city|campus|tuition|duration|language|intake)\b', candidate, re.I):
            continue
        fallback.append(candidate)

    if fallback:
        return fallback[0][:180]

    return ""


def choose_university(card: Tag, text: str, chunks: list[str], program_name: str) -> str:
    candidates = []

    for element in card.select(
        '[data-testid*="school"], [data-testid*="institution"], [data-testid*="university"], '
        '[class*="school" i], [class*="institution" i], [class*="university" i], [class*="college" i]'
    ):
        value = clean(element.get_text(" ", strip=True))
        if value:
            candidates.append(value)

    labeled = find_chunk_labeled_value(chunks, FIELD_LABELS["university"], max_len=140) or \
        find_labeled_value(text, FIELD_LABELS["university"], max_len=140)
    if labeled:
        candidates.append(labeled)

    for chunk in chunks:
        value = find_chunk_labeled_value([chunk], FIELD_LABELS["university"], max_len=140)
        candidates.append(value or chunk)

    program_normalized = normalize_text(program_name)
    seen = set()
    for candidate in candidates:
        normalized = normalize_text(candidate)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        if normalized == program_normalized:
            continue
        if looks_like_institution(candidate):
            return normalize_institution_name(candidate)[:180]

    return ""


def choose_field_from_card(card: Tag, text: str, field_name: str, max_len: int = 180) -> str:
    labels = FIELD_LABELS[field_name]
    chunks = get_card_chunks(card)

    if field_name == "city":
        return normalize_campus_city(extract_city_value(text, chunks))
    if field_name == "duration":
        return extract_duration_value(text, chunks)
    if field_name == "tuition":
        return extract_tuition_value(text, chunks)

    for label in labels:
        for element in card.select(f'[data-testid*="{label}"], [class*="{label}" i]'):
            value = clean(element.get_text(" ", strip=True))
            if value and len(value) <= max_len and is_meaningful_field_value(field_name, value, labels):
                return value

    value = find_chunk_labeled_value(chunks, labels, max_len=max_len) or \
        find_labeled_value(text, labels, max_len=max_len)
    return value if is_meaningful_field_value(field_name, value, labels) else ""


def validate_record(record: dict) -> dict:
    for key, value in list(record.items()):
        if isinstance(value, str):
            record[key] = strip_ui_noise(value)

    program_name = record.get("program_name", "")
    university = record.get("university", "")

    if program_name and looks_like_institution(program_name):
        if not university or not looks_like_institution(university):
            record["university"] = program_name
        record["program_name"] = ""

    if record.get("program_name") and record.get("university"):
        if normalize_text(record["program_name"]) == normalize_text(record["university"]):
            record["program_name"] = ""

    if not record.get("program_name"):
        fallback = find_labeled_value(record.get("raw_text", ""), ["program", "title", "name"], max_len=140)
        if fallback and not looks_like_institution(fallback):
            record["program_name"] = fallback

    return record


def merge_record(base_record: dict, detail_record: dict) -> dict:
    merged = dict(base_record)
    for key, value in detail_record.items():
        if value:
            merged[key] = value
    return validate_record(merged)


def infer_degree(value: str) -> str:
    text = normalize_text(value)
    if "foundation" in text:
        return "Foundation Programme"
    if "phd" in text or "doctor" in text:
        return "PHD"
    if "master" in text or "msc" in text or "mba" in text or re.search(r"\bma\b", text):
        return "Postgraduate"
    if "bachelor" in text or "undergraduate" in text or "bsc" in text or re.search(r"\bba\b", text):
        return "Undergraduate"
    return ""


def combine_nonempty(parts: list[str], sep: str = " | ") -> str:
    seen = []
    for part in parts:
        value = clean(part)
        if value and value not in seen:
            seen.append(value)
    return sep.join(seen)


def extract_years(*values: str) -> str:
    years = []
    for value in values:
        for year in re.findall(r"\b(20\d{2})\b", value or ""):
            if year not in years:
                years.append(year)
    return ", ".join(years)


def extract_gap_duration(*values: str) -> str:
    patterns = [
        r'(\d+(?:\.\d+)?\s*(?:years|year|yrs|yr|months|month))\s+gap',
        r'gap\s+(?:of\s+)?(\d+(?:\.\d+)?\s*(?:years|year|yrs|yr|months|month))',
        r'gap\s+accepted\s*(?:up to)?\s*(\d+(?:\.\d+)?\s*(?:years|year|yrs|yr|months|month))',
    ]
    combined = " ".join(value or "" for value in values)
    for pattern in patterns:
        match = re.search(pattern, combined, re.I)
        if match:
            return clean(match.group(1))
    return ""


def parse_money(value: str) -> tuple[str, float] | None:
    if not value:
        return None
    match = re.search(r'([A-Z]{3}|[$£€CADUSDGBP]+)?\s*([\d,]+(?:\.\d+)?)', value)
    if not match:
        return None
    currency = clean(match.group(1) or "")
    amount = float(match.group(2).replace(",", ""))
    return currency, amount


def format_money(currency: str, amount: float) -> str:
    if amount.is_integer():
        number = f"{int(amount):,}"
    else:
        number = f"{amount:,.2f}".rstrip("0").rstrip(".")
    return f"{currency} {number}".strip()


def calculate_tuition_after_scholarship(tuition_fee: str, scholarship: str) -> str:
    tuition_parsed = parse_money(tuition_fee)
    scholarship_parsed = parse_money(scholarship)
    if not tuition_parsed or not scholarship_parsed:
        return ""

    tuition_currency, tuition_amount = tuition_parsed
    scholarship_currency, scholarship_amount = scholarship_parsed

    if scholarship_amount > tuition_amount:
        return ""
    if scholarship_currency and tuition_currency and scholarship_currency != tuition_currency:
        return ""

    return format_money(tuition_currency or scholarship_currency, tuition_amount - scholarship_amount)


def extract_programme_name(program_name: str, degree_level: str = "") -> str:
    value = clean(program_name)
    if not value:
        return ""

    if " - " in value:
        left, right = [clean(part) for part in value.split(" - ", 1)]
        if looks_like_institution(left) and right:
            return right

    return value


EXPORT_COLUMNS = [
    "Country",
    "Institute",
    "Degree Type",
    "Programme Name",
    "Duration",
    "Gap Duration",
    "Intake",
    "Year",
    "Academic Requirement(s)",
    "Eng. Language Req(s)",
    "Application Fee",
    "Other Fee(s)",
    "Province/State/City",
    "Tuition Fee",
    "Scholarship",
    "Tution Fee After Scholarship",
]

EXPORT_WIDTHS = {
    "Country": 16,
    "Institute": 30,
    "Degree Type": 18,
    "Programme Name": 40,
    "Duration": 14,
    "Gap Duration": 16,
    "Intake": 22,
    "Year": 12,
    "Academic Requirement(s)": 42,
    "Eng. Language Req(s)": 30,
    "Application Fee": 18,
    "Other Fee(s)": 26,
    "Province/State/City": 24,
    "Tuition Fee": 18,
    "Scholarship": 22,
    "Tution Fee After Scholarship": 24,
}


def build_export_record(record: dict) -> dict:
    degree_type = infer_degree(
        combine_nonempty([record.get("degree_level", ""), record.get("program_name", "")], sep=" ")
    )
    programme_name = extract_programme_name(record.get("program_name", ""), record.get("degree_level", ""))
    institute = normalize_institution_name(record.get("university", ""))
    duration = extract_duration_value(record.get("duration", ""), [record.get("duration", "")])
    intakes = normalize_intake_value(record.get("detail_start_dates") or record.get("intake", ""))
    province_state_city = normalize_campus_city(record.get("detail_campus") or record.get("city", ""))
    country = record.get("country", "")
    tuition_fee = record.get("detail_tuition_detail") or record.get("tuition", "")
    scholarship = record.get("detail_scholarship", "")

    return {
        "Country": country,
        "Institute": institute,
        "Degree Type": degree_type,
        "Programme Name": programme_name,
        "Duration": duration,
        "Gap Duration": "",
        "Intake": intakes,
        "Year": extract_years(intakes, record.get("detail_deadline", "")),
        "Academic Requirement(s)": record.get("detail_academic_requirements") or record.get("detail_requirements", ""),
        "Eng. Language Req(s)": record.get("detail_english_language_requirements") or record.get("detail_english_req", ""),
        "Application Fee": record.get("detail_application_fee", ""),
        "Other Fee(s)": record.get("detail_other_fee", ""),
        "Province/State/City": province_state_city,
        "Tuition Fee": tuition_fee,
        "Scholarship": scholarship,
        "Tution Fee After Scholarship": calculate_tuition_after_scholarship(tuition_fee, scholarship),
    }


def build_export_rows(records: list[dict]) -> list[dict]:
    return [build_export_record(record) for record in records]


DETAIL_SECTION_BREAKS = {
    "Program Summary",
    "Admission Requirements",
    "Academic Background",
    "Minimum Language Test Scores",
    "Scholarships",
    "Similar Programs",
    "Program Intakes",
    "Post-Study Work Visa",
    "ApplyBoard Services",
}

LANGUAGE_TEST_LABELS = [
    "IELTS",
    "TOEFL",
    "PTE",
    "Duolingo",
    "GRE",
    "GMAT",
    "CAE",
    "CAEL",
    "MELAB",
]


def extract_nonempty_lines(text: str) -> list[str]:
    return [clean(line) for line in (text or "").splitlines() if clean(line)]


def extract_section_lines(lines: list[str], start_label: str, stop_labels: list[str] | None = None) -> list[str]:
    start_norm = normalize_text(start_label)
    stop_norms = {normalize_text(label) for label in (stop_labels or [])}
    section_breaks = {normalize_text(label) for label in DETAIL_SECTION_BREAKS}
    start_index = -1

    for index, line in enumerate(lines):
        if normalize_text(line) == start_norm:
            start_index = index + 1
            break

    if start_index < 0:
        return []

    section = []
    for line in lines[start_index:]:
        normalized = normalize_text(line)
        if normalized in stop_norms:
            break
        if section and normalized in section_breaks:
            break
        section.append(line)
    return section


def is_fee_like_value(value: str) -> bool:
    return bool(re.search(r"(?:[$£€]|[A-Z]{3}|\bFree\b|\d)", value or "", re.I))


def extract_value_after_label(
    lines: list[str],
    label: str,
    stop_labels: list[str] | None = None,
    skip_values: list[str] | None = None,
) -> str:
    label_norm = normalize_text(label)
    stop_norms = {normalize_text(item) for item in (stop_labels or [])}
    skip_norms = {normalize_text(item) for item in (skip_values or [])}

    for index, line in enumerate(lines):
        if normalize_text(line) != label_norm:
            continue
        for candidate in lines[index + 1:]:
            normalized = normalize_text(candidate)
            if normalized in stop_norms:
                return ""
            if normalized in skip_norms or not normalized:
                continue
            return candidate
    return ""


def extract_academic_requirements(lines: list[str]) -> str:
    section = extract_section_lines(
        lines,
        "Academic Background",
        stop_labels=["Minimum Language Test Scores", "Scholarships", "Similar Programs", "Program Intakes"],
    )
    requirements = []

    minimum_education = extract_value_after_label(
        section,
        "Minimum Level of Education Completed",
        stop_labels=["Minimum GPA", "Minimum Language Test Scores"],
    )
    minimum_gpa = extract_value_after_label(
        section,
        "Minimum GPA",
        stop_labels=["Minimum Language Test Scores", "Scholarships"],
        skip_values=["Convert grades"],
    )

    if minimum_education:
        requirements.append(f"Minimum Level of Education Completed: {minimum_education}")
    if minimum_gpa:
        requirements.append(f"Minimum GPA: {minimum_gpa}")

    return combine_nonempty(requirements)


def extract_language_requirements(lines: list[str]) -> str:
    section = extract_section_lines(
        lines,
        "Minimum Language Test Scores",
        stop_labels=[
            "This program requires valid language test results",
            "The program requirements above should only be used as a guide and do not guarantee admission into the program.",
            "Scholarships",
            "Similar Programs",
        ],
    )
    requirements = []
    known_labels = {normalize_text(label) for label in LANGUAGE_TEST_LABELS}

    for label in LANGUAGE_TEST_LABELS:
        value = extract_value_after_label(section, label, stop_labels=LANGUAGE_TEST_LABELS)
        if value:
            requirements.append(f"{label}: {value}")

    if requirements:
        return combine_nonempty(requirements)

    fallback = []
    for index, line in enumerate(section[:-1]):
        if normalize_text(line) in known_labels:
            value = section[index + 1]
            if value and normalize_text(value) not in known_labels:
                fallback.append(f"{line}: {value}")

    return combine_nonempty(fallback)


def extract_application_fee(lines: list[str], text: str) -> str:
    application_fee = extract_value_after_label(
        lines,
        "Application Fee",
        stop_labels=["Other Fees", "Program Intakes", "Scholarships"],
    )
    if application_fee:
        return application_fee

    match = re.search(r"Application Fee\s+([^\n]+)", text, re.I)
    return clean(match.group(1)) if match else ""


def extract_other_fee(lines: list[str], text: str) -> str:
    section = extract_section_lines(
        lines,
        "Other Fees",
        stop_labels=["Program Intakes", "Scholarships", "Similar Programs", "Post-Study Work Visa"],
    )
    fee_parts = []
    index = 0

    while index < len(section):
        label = section[index]
        if index + 1 < len(section) and is_fee_like_value(section[index + 1]):
            fee_parts.append(f"{label}: {section[index + 1]}")
            index += 2
            continue
        if label:
            fee_parts.append(label)
        index += 1

    if fee_parts:
        return combine_nonempty(fee_parts)

    match = re.search(r"Other Fees\s+(.+?)\s+Program Intakes", text, re.I | re.S)
    return clean(match.group(1))[:300] if match else ""


def parse_card(card: Tag) -> dict:
    text = card.get_text(" ", strip=True)
    chunks = get_card_chunks(card)

    link_tag = (card.select_one('a[href*="/programs/"]') or
                card.select_one('a[href*="/apply"]') or
                card.select_one('a[href^="/"]') or
                card.select_one('a[href^="http"]'))
    href = link_tag.get("href","") if link_tag else ""
    if href.startswith("/"): href = DETAIL_BASE + href

    record = {
        "program_name":  choose_program_name(card, href, text, chunks),
        "university":    "",
        "country":       choose_field_from_card(card, text, "country", max_len=80),
        "city":          choose_field_from_card(card, text, "city", max_len=80),
        "degree_level":  choose_field_from_card(card, text, "degree_level", max_len=80),
        "subject":       choose_field_from_card(card, text, "subject", max_len=120),
        "duration":      choose_field_from_card(card, text, "duration", max_len=80),
        "tuition":       choose_field_from_card(card, text, "tuition", max_len=120),
        "language":      choose_field_from_card(card, text, "language", max_len=80),
        "intake":        extract_available_intakes(text, chunks) or choose_field_from_card(card, text, "intake", max_len=120),
        "program_url":   href,
        "raw_text":      clean(text)[:600],
    }
    record["university"] = choose_university(card, text, chunks, record["program_name"])
    return validate_record(record)

def parse_all_cards(html: str, debug: bool = False) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    if debug:
        classes = sorted({c[:40] for t in soup.find_all(True) for c in (t.get("class") or [])})
        testids = [t.get("data-testid") for t in soup.find_all(attrs={"data-testid": True})]
        print(f"\n   🔬  Classes (first 60): {classes[:60]}")
        print(f"   🔬  data-testids: {testids[:40]}\n")
    cards = detect_cards(soup)
    return [r for c in cards if (r := parse_card(c)) and (r.get("program_name") or r.get("university"))]

# ──────────────────────────────────────────────────────────────────────────────
# DETAIL PAGE
# ──────────────────────────────────────────────────────────────────────────────

async def scrape_detail(page: Page, url: str, debug: bool = False, screenshot_path: Path | None = None) -> dict:
    if not url: return {}
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
        detail_ready = await wait_for_first_visible(
            page,
            [
                "h1",
                'text="Program Summary"',
                'text="Admission Requirements"',
                'text="Overview"',
            ],
            45_000,
        )
        if detail_ready:
            await settle_page(page, pause_ms=1_800)
        else:
            await settle_page(page, pause_ms=4_000)
        if screenshot_path:
            await page.screenshot(path=str(screenshot_path), full_page=True)
            print(f"      Screenshot -> {screenshot_path}")
        html = await page.content()
        if debug: save_debug_html(html, "detail")
        soup = BeautifulSoup(html, "html.parser")
        text = soup.get_text(" ", strip=True)
        body_text = await page.locator("body").inner_text()
        detail_lines = extract_nonempty_lines(body_text)
        chunks = get_card_chunks(soup, max_len=160)

        def g(*keys):
            for k in keys:
                for el in soup.select(f'[data-testid*="{k}"], [class*="{k}"]'):
                    v = clean(el.get_text())
                    if v and len(v) < 500: return v
            return ""

        def p(pattern):
            m = re.search(pattern, text, re.I)
            return clean(m.group(1))[:300] if m else ""

        meta = soup.select_one('meta[name="description"], meta[property="og:description"]')
        desc = clean(meta["content"]) if meta and meta.get("content") else g("description","overview","about")
        detail_institution, detail_program_name = extract_detail_header_fields(detail_lines)
        academic_requirements = extract_academic_requirements(detail_lines)
        english_requirements = extract_language_requirements(detail_lines)
        application_fee = extract_application_fee(detail_lines, body_text)
        other_fee = extract_other_fee(detail_lines, body_text)

        detail_record = {
            "program_name": detail_program_name or choose_program_name(soup, url, text, chunks),
            "university": detail_institution or choose_university(soup, text, chunks, detail_program_name),
            "detail_description":     desc,
            "detail_requirements":    academic_requirements or g("requirement","admission","eligibility") or p(r'(?:admission|entry)\s*requirements?[:\s]+([^.]{20,300})'),
            "detail_academic_requirements": academic_requirements,
            "detail_english_req":     english_requirements or g("english","ielts","toefl","language-req") or p(r'(?:ielts|toefl|english)[:\s]+([^.]{10,200})'),
            "detail_english_language_requirements": english_requirements,
            "detail_application_fee": application_fee or g("application-fee","applicationFee") or p(r'application\s*fee[:\s]+([^\n]{5,100})'),
            "detail_other_fee":       other_fee,
            "detail_gpa":             g("gpa","grade","average")                   or p(r'(?:gpa|grade)[:\s]+([^\n]{3,80})'),
            "detail_work_permit":     g("work-permit","coop","co-op","internship") or p(r'(?:work permit|co-?op)[:\s]+([^.]{10,200})'),
            "detail_scholarship":     g("scholarship","bursary","funding")         or p(r'scholarship[:\s]+([^.]{10,200})'),
            "detail_accreditation":   g("accreditation","accredited")              or p(r'accreditati\w+[:\s]+([^.]{10,200})'),
            "detail_campus":          normalize_campus_city(extract_city_value(text, chunks) or g("campus-location","campus-city") or p(r'(?:campus city|city|campus)[:\s]+([^\n]{2,150})')),
            "detail_start_dates":     extract_available_intakes(text, chunks, detail_lines) or normalize_intake_value(g("available-intakes","intake","start-date","start-dates")) or normalize_intake_value(p(r'available\s+intakes[:\s]+([^\n]{5,150})')) or normalize_intake_value(p(r'(?:start date|intake)[:\s]+([^\n]{5,150})')),
            "detail_deadline":        g("deadline","application-deadline")         or p(r'deadline[:\s]+([^\n]{5,100})'),
            "detail_tuition_detail":  extract_tuition_value(text, chunks) or g("tuition-detail","annual-tuition","tuition") or p(r'tuition\s*\(?(?:1st|first)\s*year\)?[:\s]+([^\n]{5,150})') or p(r'(?:annual\s+)?tuition[:\s]+([^\n]{5,150})'),
        }
        return validate_record(detail_record)
    except Exception as e:
        print(f"      ⚠  Detail error: {e}")
        return {}

# ──────────────────────────────────────────────────────────────────────────────
# EXPORT
# ──────────────────────────────────────────────────────────────────────────────

def export_xlsx(records, path, applied_filters):
    export_rows = build_export_rows(records)
    wb = Workbook()
    ws = wb.active
    ws.title = "Programs"

    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill   = PatternFill("solid", fgColor="D6E4F0")
    link_font  = Font(name="Arial", color="1155CC", underline="single", size=10)
    body_font  = Font(name="Arial", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for ci, col in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align
    ws.row_dimensions[1].height = 32

    for ri, rec in enumerate(export_rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, col in enumerate(EXPORT_COLUMNS, 1):
            val = rec.get(col, "") or ""
            cell = ws.cell(row=ri, column=ci, value=val)
            if fill: cell.fill = fill
            cell.font = body_font
            cell.alignment = wrap_align

    for ci, col in enumerate(EXPORT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = EXPORT_WIDTHS.get(col, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 20
    bold = Font(name="Arial", bold=True, size=11)
    ws2["A1"] = "ApplyBoard Scrape Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)
    ws2.append([])
    ws2.append(["Scrape Date", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws2.append(["Total Programs", len(records)])
    ws2.append([])
    if applied_filters:
        ws2.append(["── Filters Applied ──"])
        ws2[f"A{ws2.max_row}"].font = bold
        for k, v in applied_filters.items():
            ws2.append([k.replace("_"," ").title(), v])
        ws2.append([])
    df = pd.DataFrame(export_rows)
    for col_name, label in [("Country","By Country"),("Degree Type","By Degree Type"),("Institute","By Institute")]:
        if col_name in df.columns and not df[col_name].dropna().empty:
            ws2.append([f"── {label} ──", "Count"])
            r = ws2.max_row
            ws2[f"A{r}"].font = bold
            ws2[f"B{r}"].font = bold
            for val, cnt in df[col_name].value_counts().head(20).items():
                if val: ws2.append([val, cnt])
            ws2.append([])

    wb.save(path)
    print(f"\n✅  XLSX → {path}")

def export_csv(records, path):
    export_rows = build_export_rows(records)
    pd.DataFrame(export_rows, columns=EXPORT_COLUMNS).to_csv(path, index=False, encoding="utf-8-sig")
    print(f"✅  CSV  → {path}")

# ──────────────────────────────────────────────────────────────────────────────
# PROMPT
# ──────────────────────────────────────────────────────────────────────────────

def prompt_user(force_login: bool = False):
    print("\n" + "═"*60)
    print("  ApplyBoard Program Scraper  🎓")
    print("═"*60)

    print("\n📌  Filters (press Enter to skip any)\n")
    filters = {}
    val = input("  University name (e.g. University of Toronto): ").strip()
    if val: filters["university"] = val

    val = input("  Country (e.g. Canada, UK, Australia): ").strip()
    if val: filters["country"] = val

    val = input("  Subject (e.g. Computer Science, Business): ").strip()
    if val: filters["subject"] = val

    val = input("  Degree Level (e.g. Bachelor, Master, Diploma): ").strip()
    if val: filters["degree_level"] = val

    print()
    do_details = input("🔍  Scrape detail pages? (y/n) [y]: ").strip().lower() != "n"
    use_login  = force_login or input("Login to ApplyBoard first? (y/n) [n]: ").strip().lower() == "y"

    auth = {"enabled": use_login, "email": "", "password": ""}
    if use_login:
        env_email = os.getenv("APPLYBOARD_EMAIL", "").strip()
        env_password = os.getenv("APPLYBOARD_PASSWORD", "").strip()
        email_prompt = "  ApplyBoard email"
        if env_email:
            email_prompt += f" [{mask_email(env_email)}]"
        email = input(f"{email_prompt}: ").strip() or env_email
        password = getpass("  ApplyBoard password: ").strip() or env_password
        if not email or not password:
            raise ValueError(
                "Login was enabled, but the ApplyBoard email/password was not provided."
            )
        auth = {"enabled": True, "email": email, "password": password}

    screenshot_count = 0
    screenshot_dir = ""
    if do_details:
        shot_input = input("Save screenshots of the first N detail pages? [0]: ").strip()
        screenshot_count = int(shot_input) if shot_input.isdigit() else 0
        if screenshot_count > 0:
            screenshot_dir = (
                input("Screenshot folder [program_detail_screenshots]: ").strip()
                or "program_detail_screenshots"
            )
    max_p      = input("📄  Max pages (0 = all) [0]: ").strip()
    max_pages  = int(max_p) if max_p.isdigit() else 0
    out_fmt    = input("💾  Output: xlsx / csv / both [xlsx]: ").strip().lower() or "xlsx"
    out_name   = input("📁  Filename (no extension) [applyboard_programs]: ").strip() or "applyboard_programs"

    return filters, do_details, max_pages, out_fmt, out_name, auth, screenshot_count, screenshot_dir

# ──────────────────────────────────────────────────────────────────────────────
# MAIN RUN
# ──────────────────────────────────────────────────────────────────────────────

async def run(filters, do_details, max_pages, out_fmt, out_name,
              auth=None, screenshot_count=0, screenshot_dir="",
              debug=False, headless=False):

    all_records = []
    seen_program_keys = set()
    auth = auth or {"enabled": False, "email": "", "password": ""}
    screenshot_root = Path(screenshot_dir) if screenshot_dir else None
    storage_state_path = SESSION_STATE_PATH if has_saved_session() else None
    if screenshot_count > 0 and screenshot_root:
        screenshot_root.mkdir(parents=True, exist_ok=True)

    async with async_playwright() as pw:
        browser, ctx = await create_browser(
            pw,
            headless=headless,
            storage_state_path=storage_state_path,
        )
        page = await ctx.new_page()

        if auth.get("enabled"):
            session_active = False
            if storage_state_path:
                print(f"\n[session] Reusing saved ApplyBoard session -> {storage_state_path}")
                session_active = await has_active_agent_session(page, debug=debug)
            if not session_active:
                await login_to_applyboard(page, auth.get("email", ""), auth.get("password", ""), debug=debug)
                await ctx.storage_state(path=str(SESSION_STATE_PATH))
                print(f"   [ok] Saved session -> {SESSION_STATE_PATH}")
        elif storage_state_path:
            print(f"\n[session] Loaded saved ApplyBoard session -> {storage_state_path}")

        print(f"\n🌐  Opening search page → {BASE_URL}")
        await page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30_000)
        await settle_page(page, pause_ms=2_400)
        if debug:
            save_debug_html(await page.content(), "search_landing")

        # Step 1: Apply the university filter in the search UI if provided
        filtered_base_url = page.url
        if filters.get("university"):
            filtered_base_url = await apply_university_filter(page, filters["university"], debug=debug)
            if not filtered_base_url:
                print(f"\n❌  Could not apply the university filter for '{filters['university']}'.")
                print("   Stopping here so the scraper does not continue with unfiltered results.")
                if debug:
                    save_debug_html(await page.content(), "university_filter_not_resolved")
                await browser.close()
                return []

        filters_ready = await wait_for_filters_ready(page, filters, filtered_base_url, debug=debug)
        if not filters_ready:
            print("\n❌  Filters checkpoint failed. Scraping was stopped before any results were collected.")
            await browser.close()
            return []

        # Step 2: Paginate using direct URL construction
        detail_tab = await ctx.new_page() if do_details else None
        page_num   = 1

        while True:
            url = build_search_url(
                page_number  = page_num,
                base_url     = filtered_base_url,
                country      = filters.get("country"),
                subject      = filters.get("subject"),
                degree_level = filters.get("degree_level"),
            )

            print(f"\n📄  Page {page_num} → {url}")
            if page_num == 1:
                await settle_page(page, pause_ms=max(2_000, PAGE_LOAD_WAIT // 4))
            else:
                await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
                await settle_page(page, pause_ms=max(2_000, PAGE_LOAD_WAIT // 4))
            await human_scroll(page)
            await human_pause(900, 1600)

            html = await page.content()
            if debug:
                save_debug_html(html, f"page_{page_num}")

            records = parse_all_cards(html, debug=(debug and page_num == 1))
            print(f"   Found {len(records)} programs")

            if not records:
                print("   ✅  No more results.")
                break

            new_page_records = []
            repeated_page_records = 0
            for rec in records:
                program_url = rec.get("program_url", "")
                record_key = (
                    program_url,
                    normalize_text(rec.get("program_name", "")),
                    normalize_text(rec.get("university", "")),
                    normalize_intake_value(rec.get("intake", "")),
                )
                if record_key in seen_program_keys:
                    repeated_page_records += 1
                    continue
                seen_program_keys.add(record_key)
                new_page_records.append(rec)

            if not new_page_records:
                print("   ✅  No new programs found on this page. Stopping pagination.")
                break

            if repeated_page_records:
                print(f"   Skipped {repeated_page_records} programs already captured from earlier pages")

            records = new_page_records

            if do_details and detail_tab:
                for i, rec in enumerate(records):
                    url_d = rec.get("program_url","")
                    if url_d and url_d.startswith("http"):
                        print(f"   [{i+1}/{len(records)}] {url_d[:80]}")
                        screenshot_path = None
                        should_capture = screenshot_root and len(all_records) + i < screenshot_count
                        if should_capture:
                            program_stub = sanitize_filename(
                                rec.get("program_name") or rec.get("university") or "program"
                            )
                            screenshot_path = screenshot_root / f"page_{page_num:02d}_{i+1:02d}_{program_stub}.png"
                        detail_data = await scrape_detail(
                            detail_tab,
                            url_d,
                            debug=(debug and i == 0),
                            screenshot_path=screenshot_path,
                        )
                        if detail_data:
                            rec = merge_record(rec, detail_data)
                            records[i] = rec
                        await human_pause(900, 2000)

            all_records.extend(records)

            if max_pages and page_num >= max_pages:
                print("   ✋  Page limit reached.")
                break

            page_num += 1
            await human_pause(1500, 3200)

        if detail_tab: await detail_tab.close()
        await browser.close()

    # Deduplicate
    seen, unique = set(), []
    for r in all_records:
        key = (r.get("program_name","").lower(), r.get("university","").lower(), r.get("program_url",""))
        if key not in seen:
            seen.add(key)
            unique.append(r)

    print(f"\n📊  Total unique programs: {len(unique)}")

    if not unique:
        print("\n⚠  No data scraped.")
        print("   Run with --debug to save HTML and inspect the page structure.")
        return unique

    if out_fmt in ("xlsx","both"): export_xlsx(unique, f"{out_name}.xlsx", filters)
    if out_fmt in ("csv","both"):  export_csv(unique,  f"{out_name}.csv")
    return unique


def main():
    parser = argparse.ArgumentParser(description="ApplyBoard scraper")
    parser.add_argument("--debug",      action="store_true", help="Save raw HTML for inspection")
    parser.add_argument("--login",      action="store_true", help="Prompt for ApplyBoard login before scraping")
    parser.add_argument("--no-details", action="store_true", help="Skip detail pages")
    parser.add_argument("--headless",   action="store_true", help="Headless browser mode")
    parser.add_argument("--screenshots", type=int, default=0, help="Save the first N detail pages as screenshots")
    parser.add_argument("--screenshot-dir", default="", help="Folder for detail page screenshots")
    args = parser.parse_args()

    filters, do_details, max_pages, out_fmt, out_name, auth, screenshot_count, screenshot_dir = prompt_user(
        force_login=args.login
    )
    if args.login:
        auth["enabled"] = True
        if not auth.get("email"):
            auth["email"] = os.getenv("APPLYBOARD_EMAIL", "").strip()
        if not auth.get("password"):
            auth["password"] = os.getenv("APPLYBOARD_PASSWORD", "").strip()
        if not auth.get("email") or not auth.get("password"):
            raise ValueError(
                "--login was provided, but ApplyBoard credentials were not supplied in the prompt "
                "or via APPLYBOARD_EMAIL / APPLYBOARD_PASSWORD."
            )
    if args.no_details: do_details = False
    if args.screenshots:
        screenshot_count = args.screenshots
    if args.screenshot_dir:
        screenshot_dir = args.screenshot_dir
    if screenshot_count > 0 and not args.no_details:
        do_details = True

    asyncio.run(run(filters, do_details, max_pages,
                    out_fmt, out_name, auth=auth,
                    screenshot_count=screenshot_count, screenshot_dir=screenshot_dir,
                    debug=args.debug, headless=args.headless))


if __name__ == "__main__":
    main()
